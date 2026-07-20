"""Tests for the console/XLSX formatters (veta.output)."""

from __future__ import annotations

import datetime

from veta import output
from veta.intelligence import BuyerIntel, EnrichedTender, Urgency


def _tender(**overrides):
    base = dict(
        numero_procedimiento="LA-1",
        nombre_procedimiento="Compra de prueba",
        siglas="IMSS",
        tipo_contratacion="ADQUISICIONES",
        caracter="NACIONAL",
        estatus_alterno="VIGENTE",
        entidad_federativa="CDMX",
        unidad_compradora="UC1",
        uuid_procedimiento="uuid-1",
        matched_partidas=["25401"],
        intel=[BuyerIntel("IMSS", "25401", "med", True, contract_count=5, new_entrant_rate=0.4)],
        urgency=Urgency(datetime.datetime(2026, 1, 20, 9, 0), 10, None, None, "GREEN"),
        signal="STRONG: open buyer",
    )
    base.update(overrides)
    return EnrichedTender(**base)


def test_money():
    assert output._money(None) == "n/a"
    assert output._money(1234567) == "$1,234,567 MXN"


def test_month_name():
    assert output._month_name(9) == "Sep"
    assert output._month_name(None) == "n/a"
    assert output._month_name(13) == "n/a"


def test_monto_line_always_shown_not_published():
    t = _tender(line_partidas=None, monto_min=None, monto_max=None)
    assert output._monto_line(t) == "Est. value: not published by buyer"


def test_monto_line_not_published():
    t = _tender(line_partidas=1, monto_min=None, monto_max=None)
    assert output._monto_line(t) == "Est. value: not published by buyer"


def test_monto_line_band():
    t = _tender(line_partidas=1, monto_min=100.0, monto_max=200.0)
    assert output._monto_line(t) == "Est. value: $100 MXN to $200 MXN"


def test_render_card_contains_key_fields():
    card = output.render_card(_tender())
    assert "LA-1" in card
    assert "SIGNAL:" in card
    assert "Urgency:   GREEN" in card


def test_to_dataframe_has_monto_columns():
    frame = output.to_dataframe([_tender(line_partidas=1, monto_min=100.0, monto_max=200.0)])
    assert "est_monto_min" in frame.columns
    assert "est_monto_max" in frame.columns
    assert frame.iloc[0]["est_monto_min"] == 100.0
    assert frame.iloc[0]["numero_procedimiento"] == "LA-1"


def test_senal_es_translates_grades():
    assert output._senal_es("STRONG: x") == "FUERTE"
    assert output._senal_es("MODERATE: x") == "MODERADA"
    assert output._senal_es("WEAK: x") == "DEBIL"
    assert output._senal_es("NO HISTORY") == "SIN HISTORIAL"
    assert output._senal_es("UNVERIFIED MATCH: x") == "SIN VERIFICAR"


def test_monto_cell_not_published_and_band():
    assert output._monto_cell(_tender(monto_min=None, monto_max=None)) == "No publicado"
    assert output._monto_cell(_tender(monto_min=100.0, monto_max=200.0)) == "$100 - $200 MXN"


def test_default_report_path_is_dated():
    import datetime

    path = output.default_report_path()
    today = datetime.date.today().strftime("%Y-%m-%d")
    assert path == f"reports/reporte-veta-{today}.xlsx"


def test_write_client_xlsx_has_two_named_sheets(tmp_path):
    from openpyxl import load_workbook

    path = str(tmp_path / "reporte.xlsx")
    t = _tender(line_partidas=50, monto_min=100.0, monto_max=200.0)
    output.write_client_xlsx([t], path)

    wb = load_workbook(path)
    assert wb.sheetnames == ["Resumen", "Detalle"]
    resumen = wb["Resumen"]
    assert resumen["A1"].value == "VETA - Reporte de Inteligencia"
    assert resumen["A6"].value == "Accion"
    # Header row 6, first data row 7 carries the tender.
    assert resumen["B7"].value == "LA-1"
    detalle = wb["Detalle"]
    assert detalle["A1"].value == "Accion"
    assert detalle["B1"].value == "No. Procedimiento"


def _position(grade, plow=0.20, phigh=0.30, prior=0):
    from veta.positioning import ClientPosition

    return ClientPosition(
        position_grade=grade, position_score=0.5, p_win_low=plow, p_win_high=phigh,
        n_prior_wins=prior, last_win_year=2025, share_of_buyer=0.1,
        n_other_buyers_same_partida=2, n_other_partidas_same_buyer=3,
        incumbency_strength=0.5, category_strength=0.4, relationship_strength=0.3,
    )


def _positioned_tender(pos_grade, base_grade="STRONG", num="LA-1", deadline=None):
    intel = BuyerIntel(
        "IMSS", "25401", "med", True, contract_count=5, new_entrant_rate=0.4,
        base_grade=base_grade, confidence="high",
    )
    intel.position = _position(pos_grade)
    dl = deadline or datetime.datetime(2026, 1, 20, 9, 0)
    return _tender(
        numero_procedimiento=num, intel=[intel],
        urgency=Urgency(dl, 10, None, None, "GREEN"),
    )


def _header_row(ws):
    return [c.value for c in ws[6]]


def test_client_xlsx_uses_strategic_buckets_when_positioned(tmp_path):
    from openpyxl import load_workbook

    path = str(tmp_path / "reporte.xlsx")
    output.write_client_xlsx([_positioned_tender("EXPERIENCED", "STRONG")], path)

    ws = load_workbook(path)["Resumen"]
    assert ws["A4"].value.startswith("OPORTUNIDAD")
    headers = _header_row(ws)
    assert "Posicion" in headers
    assert "P Estimada" in headers
    assert "Contratos Previos" in headers
    assert ws["A7"].value == "OPORTUNIDAD"
    pos_col = headers.index("Posicion") + 1
    assert ws.cell(row=7, column=pos_col).value == "CON EXPERIENCIA"


def test_client_xlsx_uses_urgency_buckets_when_no_rfc(tmp_path):
    from openpyxl import load_workbook

    from veta import intelligence

    path = str(tmp_path / "reporte.xlsx")
    output.write_client_xlsx([_tender()], path)

    ws = load_workbook(path)["Resumen"]
    assert ws["A4"].value.startswith("ACTUAR")
    headers = _header_row(ws)
    assert "Posicion" not in headers
    assert ws["A7"].value in intelligence.BUCKETS


def test_oportunidad_sorts_experienced_before_adjacent(tmp_path):
    from openpyxl import load_workbook

    # ADJACENT closes sooner, but EXPERIENCED must still rank first within
    # OPORTUNIDAD (category expertise outranks buyer relationship alone).
    exp = _positioned_tender(
        "EXPERIENCED", "STRONG", num="EXP",
        deadline=datetime.datetime(2026, 2, 1, 9, 0),
    )
    adj = _positioned_tender(
        "ADJACENT", "STRONG", num="ADJ",
        deadline=datetime.datetime(2026, 1, 15, 9, 0),
    )
    path = str(tmp_path / "reporte.xlsx")
    output.write_client_xlsx([adj, exp], path)

    ws = load_workbook(path)["Resumen"]
    assert ws["B7"].value == "EXP"
    assert ws["B8"].value == "ADJ"
