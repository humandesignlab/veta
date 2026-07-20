"""Output formatters (spec section 3.5, step 5).

Renders the enriched shortlist (from veta.intelligence) as console intelligence
cards and as an XLSX workbook with one row per tender. No em dashes in any
output; monetary values are Mexican pesos.
"""

from __future__ import annotations

import datetime
import json

import pandas as pd

from veta import filters, intelligence
from veta.intelligence import BuyerIntel, EnrichedTender

MONTHS = [
    "", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

CARD_WIDTH = 60

# Spanish labels for the signal grades used in the client report.
SIGNAL_ES = {
    "STRONG": "FUERTE",
    "MODERATE": "MODERADA",
    "WEAK": "DEBIL",
    "NO HISTORY": "SIN HISTORIAL",
    "UNVERIFIED MATCH": "SIN VERIFICAR",
}


def _money(value: float | None) -> str:
    if value is None:
        return "n/a"
    return f"${value:,.0f} MXN"


def _month_name(month: int | None) -> str:
    if month is None or month < 1 or month > 12:
        return "n/a"
    return MONTHS[month]


def _deadline_line(tender: EnrichedTender) -> str:
    u = tender.urgency
    if u.deadline is None:
        return "Deadline:  unknown"
    days = u.days_to_deadline
    if days is None:
        remaining = ""
    elif days < 0:
        remaining = f" ({abs(days)} days ago)"
    else:
        remaining = f" ({days} days remaining)"
    return f"Deadline:  {u.deadline.date()}{remaining}"


def _aclaraciones_line(tender: EnrichedTender) -> str:
    u = tender.urgency
    if u.aclaraciones is None:
        return "Clarif.:   none listed"
    state = "PASSED" if u.aclaraciones_passed else "upcoming"
    return f"Clarif.:   {u.aclaraciones.date()} ({state})"


def _monto_line(tender: EnrichedTender) -> str:
    """Estimated value line. Always shown; buyers rarely publish amounts."""
    if tender.monto_min is None and tender.monto_max is None:
        return "Est. value: not published by buyer"
    if tender.monto_min == tender.monto_max:
        return f"Est. value: {_money(tender.monto_min)}"
    return f"Est. value: {_money(tender.monto_min)} to {_money(tender.monto_max)}"


def _intel_block(intel: BuyerIntel) -> list[str]:
    if not intel.has_history:
        return [
            f"BUYER INTELLIGENCE ({intel.siglas} + partida {intel.partida}):",
            "  No federal history for this buyer and category in 2023-2025.",
        ]
    rate = intel.new_entrant_rate
    open_flag = "  [OPEN BUYER]" if intel.is_open_buyer else ""
    lines = [
        f"BUYER INTELLIGENCE ({intel.siglas} + partida {intel.partida}, 2023-2025):",
        f"  Category:            {intel.partida_desc}",
        f"  Contracts awarded:   {intel.contract_count:,}",
        f"  Distinct suppliers:  {intel.distinct_suppliers:,}",
        f"  New entrant rate:    {rate:.0%}{open_flag}" if rate is not None else "  New entrant rate:    n/a",
        f"  Price band (P10-P90): {_money(intel.price_p10)} to {_money(intel.price_p90)}",
        f"  Median value:        {_money(intel.price_median)}",
        f"  Recurrence:          {'recurring ' + str(intel.years_active) if intel.is_recurring else 'irregular ' + str(intel.years_active)}",
        f"  Typical month:       {_month_name(intel.typical_month)}",
    ]
    if intel.top_suppliers:
        lines.append("  Top winners:")
        for s in intel.top_suppliers[:3]:
            lines.append(
                f"    {s['proveedor'][:44]} "
                f"({s['count']} contracts, {_money(s['total'])})"
            )
    return lines


def _position_block(intel: BuyerIntel) -> list[str]:
    pos = intel.position
    if pos is None:
        return []
    band = f"{pos.p_win_low:.0%} to {pos.p_win_high:.0%}"
    lines = [
        "",
        f"YOUR POSITION ({pos.position_grade}):",
        f"  Est. win probability: {band} (estimate, not a prediction)",
        f"  Prior wins here:      {pos.n_prior_wins}"
        + (f" (last {pos.last_win_year})" if pos.last_win_year else ""),
        f"  Share of this buyer:  {pos.share_of_buyer:.0%}",
        f"  Same partida elsewhere: {pos.n_other_buyers_same_partida} other buyers",
        f"  Other partidas here:    {pos.n_other_partidas_same_buyer} categories",
    ]
    return lines


def render_card(tender: EnrichedTender) -> str:
    bar = "=" * CARD_WIDTH
    lines = [
        bar,
        tender.numero_procedimiento,
        tender.nombre_procedimiento[:CARD_WIDTH] or "(no title)",
        bar,
        f"Buyer:     {tender.siglas} ({tender.entidad_federativa})",
        f"Type:      {tender.tipo_contratacion} | {tender.caracter}",
        _deadline_line(tender),
        _aclaraciones_line(tender),
        f"Urgency:   {tender.urgency.level}",
        _monto_line(tender),
        "",
    ]
    primary = tender.primary_intel
    if primary is not None:
        lines.extend(_intel_block(primary))
        if tender.line_partidas:
            matched = tender.line_item_counts.get(primary.partida, 0)
            share = f"  Line items in this category: {matched} of {tender.line_partidas}"
            if matched and matched * 2 < tender.line_partidas:
                share += "  [minority line]"
            lines.append(share)
        lines.extend(_position_block(primary))
    lines.append("")
    lines.append(f"SIGNAL: {tender.signal}")
    lines.append(bar)
    return "\n".join(lines)


def render_console(shortlist: list[EnrichedTender]) -> str:
    header = (
        f"VETA SHORTLIST: {len(shortlist)} active licitaciones publicas "
        "with buyer intelligence\n"
    )
    positioned = any(
        t.primary_intel is not None and t.primary_intel.position is not None
        for t in shortlist
    )
    if positioned:
        c = {b: 0 for b in intelligence.STRATEGIC_BUCKETS}
        for t in shortlist:
            c[intelligence.assign_strategic_bucket(t)] += 1
        header += (
            f"Resumen: {c['OPORTUNIDAD']} oportunidades nuevas, "
            f"{c['TERRITORIO']} en tu territorio, {c['EXPLORAR']} para explorar, "
            f"{c['NO PRIORITARIO']} no prioritarios\n"
        )
    return header + "\n\n".join(render_card(t) for t in shortlist)


def to_dataframe(shortlist: list[EnrichedTender]) -> pd.DataFrame:
    """One row per tender with intelligence columns for XLSX export."""
    rows = []
    for t in shortlist:
        p = t.primary_intel
        pos = p.position if p else None
        rows.append(
            {
                "numero_procedimiento": t.numero_procedimiento,
                "nombre_procedimiento": t.nombre_procedimiento,
                "siglas": t.siglas,
                "entidad_federativa": t.entidad_federativa,
                "unidad_compradora": t.unidad_compradora,
                "tipo_contratacion": t.tipo_contratacion,
                "caracter": t.caracter,
                "estatus": t.estatus_alterno,
                "matched_partidas": ", ".join(t.matched_partidas),
                "deadline": t.urgency.deadline.date() if t.urgency.deadline else None,
                "days_to_deadline": t.urgency.days_to_deadline,
                "aclaraciones_passed": t.urgency.aclaraciones_passed,
                "urgency": t.urgency.level,
                "est_monto_min": t.monto_min,
                "est_monto_max": t.monto_max,
                "signal": t.signal,
                "hist_partida": p.partida if p else None,
                "hist_category": p.partida_desc if p else None,
                "hist_contract_count": p.contract_count if p else None,
                "hist_distinct_suppliers": p.distinct_suppliers if p else None,
                "hist_new_entrant_rate": p.new_entrant_rate if p else None,
                "hist_price_p10": p.price_p10 if p else None,
                "hist_price_median": p.price_median if p else None,
                "hist_price_p90": p.price_p90 if p else None,
                "hist_recurring": p.is_recurring if p else None,
                "hist_typical_month": p.typical_month if p else None,
                "hist_top_suppliers": json.dumps(p.top_suppliers, ensure_ascii=False) if p else None,
                "hist_hhi": p.hhi if p else None,
                "hist_openness_shrunk": p.openness_shrunk if p else None,
                "hist_value_pctile": p.value_pctile if p else None,
                "hist_contestability": p.contestability_score if p else None,
                "hist_confidence": p.confidence if p else None,
                "hist_grade": p.base_grade if p else None,
                "pos_grade": pos.position_grade if pos else None,
                "pos_p_low": pos.p_win_low if pos else None,
                "pos_p_high": pos.p_win_high if pos else None,
                "pos_prior_wins": pos.n_prior_wins if pos else None,
                "pos_last_win_year": pos.last_win_year if pos else None,
                "pos_share_of_buyer": pos.share_of_buyer if pos else None,
                "pos_score": pos.position_score if pos else None,
                "pos_incumbency": pos.incumbency_strength if pos else None,
                "pos_category": pos.category_strength if pos else None,
                "pos_relationship": pos.relationship_strength if pos else None,
                "pos_other_buyers": pos.n_other_buyers_same_partida if pos else None,
                "pos_other_partidas": pos.n_other_partidas_same_buyer if pos else None,
                "uuid_procedimiento": t.uuid_procedimiento,
            }
        )
    return pd.DataFrame(rows)


def check_xlsx_writable(path: str) -> None:
    """Raise a clear error if an XLSX cannot be written to path.

    Checks that openpyxl is importable in the active environment and that the
    parent directory exists. Meant to be called before a long fetch so a
    misconfigured run fails fast instead of after minutes of work.
    """
    import os

    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to write XLSX files but is not installed in "
            "the active Python environment. Activate the project venv "
            "(source .venv/bin/activate) or run pip install -r requirements.txt."
        ) from exc
    parent = os.path.dirname(os.path.abspath(path))
    if not os.path.isdir(parent):
        raise RuntimeError(f"output directory does not exist: {parent}")


def write_raw_xlsx(shortlist: list[EnrichedTender], path: str) -> str:
    """Write the raw data export (English headers, one sheet). Internal use."""
    check_xlsx_writable(path)
    frame = to_dataframe(shortlist)
    frame.to_excel(path, index=False, engine="openpyxl", sheet_name="shortlist")
    return path


# --------------------------------------------------------------------------- #
# Client-facing Spanish report (two sheets: Resumen + Detalle).
# --------------------------------------------------------------------------- #

# Action-bucket cell styling: (background hex, font hex, bold). Covers both the
# urgency buckets (no CLIENT_RFC) and the strategic buckets (CLIENT_RFC set).
_BUCKET_STYLE = {
    "ACTUAR": ("FF4444", "FFFFFF", True),
    "PREPARAR": ("FFAA00", "000000", False),
    "MONITOREAR": ("44AA44", "FFFFFF", True),
    "DESCARTAR": ("CCCCCC", "000000", False),
    "OPORTUNIDAD": ("2563EB", "FFFFFF", True),
    "TERRITORIO": ("16A34A", "FFFFFF", True),
    "EXPLORAR": ("D97706", "000000", False),
    "NO PRIORITARIO": ("9CA3AF", "000000", False),
}

# Position grade -> Spanish label for the client report.
POSITION_ES = {
    "INCUMBENT": "TITULAR",
    "EXPERIENCED": "CON EXPERIENCIA",
    "ADJACENT": "ADYACENTE",
    "OUTSIDER": "NUEVO",
}

# Signal cell styling reuses the same palette by severity.
_SIGNAL_STYLE = {
    "FUERTE": ("44AA44", "FFFFFF", True),
    "MODERADA": ("FFAA00", "000000", False),
    "DEBIL": ("FF4444", "FFFFFF", True),
    "SIN HISTORIAL": ("CCCCCC", "000000", False),
    "SIN VERIFICAR": ("CCCCCC", "000000", False),
}

_RESUMEN_HEADERS = [
    "Accion", "No. Procedimiento", "Institucion", "Estado", "Descripcion",
    "Categoria", "Fecha Apertura", "Dias Restantes (cal.)", "Monto Estimado",
    "Banda Historica", "Mediana Hist.", "Tasa Nuevos", "Señal", "Items",
    "Competidores",
]
_RESUMEN_WIDTHS = [12, 32, 10, 16, 50, 40, 12, 18, 18, 24, 16, 8, 14, 10, 50]

# (dataframe key, Spanish header) for the Detalle sheet, in column order.
_DETALLE_COLUMNS = [
    ("numero_procedimiento", "No. Procedimiento"),
    ("nombre_procedimiento", "Descripcion"),
    ("siglas", "Institucion"),
    ("entidad_federativa", "Estado"),
    ("unidad_compradora", "Unidad Compradora"),
    ("tipo_contratacion", "Tipo Contratacion"),
    ("caracter", "Caracter"),
    ("estatus", "Estatus"),
    ("matched_partidas", "Partidas Coincidentes"),
    ("deadline", "Fecha Apertura"),
    ("days_to_deadline", "Dias Restantes (cal.)"),
    ("aclaraciones_passed", "Aclaraciones Pasadas"),
    ("urgency", "Urgencia"),
    ("est_monto_min", "Monto Minimo Est."),
    ("est_monto_max", "Monto Maximo Est."),
    ("signal", "Senal"),
    ("hist_partida", "Partida Historica"),
    ("hist_category", "Categoria Historica"),
    ("hist_contract_count", "Contratos Historicos"),
    ("hist_distinct_suppliers", "Proveedores Distintos"),
    ("hist_new_entrant_rate", "Tasa Nuevos Entrantes"),
    ("hist_price_p10", "Precio P10"),
    ("hist_price_median", "Precio Mediana"),
    ("hist_price_p90", "Precio P90"),
    ("hist_recurring", "Recurrente"),
    ("hist_typical_month", "Mes Tipico"),
    ("hist_hhi", "HHI"),
    ("hist_openness_shrunk", "Apertura Ajustada"),
    ("hist_value_pctile", "Percentil Valor"),
    ("hist_contestability", "Puntaje Mercado"),
    ("hist_confidence", "Confianza"),
    ("hist_grade", "Grado Mercado"),
    ("pos_grade", "Posicion"),
    ("pos_p_low", "P Min"),
    ("pos_p_high", "P Max"),
    ("pos_prior_wins", "Contratos Previos"),
    ("pos_last_win_year", "Ultimo Año Ganado"),
    ("pos_share_of_buyer", "Participacion Comprador"),
    ("pos_score", "Puntaje Posicion"),
    ("pos_incumbency", "Fuerza Incumbencia"),
    ("pos_category", "Fuerza Categoria"),
    ("pos_relationship", "Fuerza Relacion"),
    ("pos_other_buyers", "Otros Compradores"),
    ("pos_other_partidas", "Otras Partidas"),
    ("hist_top_suppliers", "Top Proveedores"),
    ("uuid_procedimiento", "UUID"),
]
# Detalle keys that hold peso amounts (number format "#,##0").
_DETALLE_MONEY = {
    "est_monto_min", "est_monto_max",
    "hist_price_p10", "hist_price_median", "hist_price_p90",
}
# Detalle keys shown as percentages (0%).
_DETALLE_PERCENT = {
    "hist_new_entrant_rate", "hist_openness_shrunk", "hist_value_pctile",
    "pos_p_low", "pos_p_high", "pos_share_of_buyer",
}
# Detalle keys shown as 2-decimal ratios (0.00).
_DETALLE_RATIO = {
    "hist_hhi", "hist_contestability", "pos_score",
    "pos_incumbency", "pos_category", "pos_relationship",
}


def _senal_es(signal: str) -> str:
    return SIGNAL_ES.get(intelligence.signal_grade(signal), "SIN HISTORIAL")


def _monto_cell(t: EnrichedTender) -> str:
    if t.monto_min is None and t.monto_max is None:
        return "No publicado"
    if t.monto_min == t.monto_max:
        return f"${(t.monto_min or 0):,.0f} MXN"
    return f"${(t.monto_min or 0):,.0f} - ${(t.monto_max or 0):,.0f} MXN"


def _banda_cell(p: BuyerIntel | None) -> str:
    if p is None or not p.has_history or p.price_p10 is None:
        return "Sin historial"
    return f"${p.price_p10:,.0f} - ${p.price_p90:,.0f} MXN"


def _categoria_cell(t: EnrichedTender) -> str:
    p = t.primary_intel
    if p is None:
        return ", ".join(t.matched_partidas)
    return f"{p.partida} - {p.partida_desc}"[:60]


def _items_cell(t: EnrichedTender) -> str:
    p = t.primary_intel
    if p is None or not t.line_partidas:
        return "S/D"
    return f"{t.line_item_counts.get(p.partida, 0)} de {t.line_partidas}"


def _competidores_cell(p: BuyerIntel | None) -> str:
    if p is None or not p.top_suppliers:
        return ""
    names = [str(s.get("proveedor", ""))[:28] for s in p.top_suppliers[:3]]
    return ", ".join(n for n in names if n)[:120]


def _sort_key(t: EnrichedTender) -> tuple[int, datetime.date]:
    bucket = intelligence.assign_bucket(t)
    bucket_rank = intelligence.BUCKETS.index(bucket)
    deadline = t.urgency.deadline.date() if t.urgency.deadline else datetime.date.max
    return (bucket_rank, deadline)


def _strategic_sort_key(t: EnrichedTender) -> tuple[int, int, datetime.date]:
    """Strategic sort: bucket order, then EXPERIENCED before ADJACENT within
    OPORTUNIDAD (category expertise outranks buyer relationship alone), then
    deadline ascending so time-sensitive tenders still surface within a group.
    """
    bucket = intelligence.assign_strategic_bucket(t)
    bucket_rank = intelligence.STRATEGIC_BUCKETS.index(bucket)
    sub = 0
    if bucket == "OPORTUNIDAD":
        p = t.primary_intel
        grade = p.position.position_grade if (p and p.position) else ""
        sub = {"EXPERIENCED": 0, "ADJACENT": 1}.get(grade, 2)
    deadline = t.urgency.deadline.date() if t.urgency.deadline else datetime.date.max
    return (bucket_rank, sub, deadline)


def default_report_path() -> str:
    """Default client-report filename, dated for the day it is run."""
    return f"reports/reporte-veta-{datetime.date.today():%Y-%m-%d}.xlsx"


def write_client_xlsx(shortlist: list[EnrichedTender], path: str) -> str:
    """Write the client-facing Spanish report (Resumen + Detalle). Returns path."""
    check_xlsx_writable(path)
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # With a client RFC, positioning is active: lead with strategic value
    # (opportunities the client is missing) instead of urgency. Otherwise keep
    # the urgency buckets - market-level intelligence with no client position.
    positioned = any(
        t.primary_intel is not None and t.primary_intel.position is not None
        for t in shortlist
    )
    if positioned:
        bucket_of = intelligence.assign_strategic_bucket
        buckets = intelligence.STRATEGIC_BUCKETS
        ordered = sorted(shortlist, key=_strategic_sort_key)
    else:
        bucket_of = intelligence.assign_bucket
        buckets = intelligence.BUCKETS
        ordered = sorted(shortlist, key=_sort_key)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    def _fill(bg: str) -> PatternFill:
        return PatternFill("solid", fgColor=bg)

    wb = Workbook()

    # ---- Sheet 1: Resumen ------------------------------------------------- #
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "VETA - Reporte de Inteligencia"
    ws["A1"].font = Font(bold=True, size=14)
    today = datetime.date.today().strftime("%d/%m/%Y")
    ncat = len(filters.INCLUDE_PARTIDAS)
    if positioned and filters.CLIENT_RFC:
        ws["A2"] = (
            f"Cliente: {filters.CLIENT_RFC} | Generado: {today} | Perfil: "
            f"LAASSP, {ncat} categorias | Ordenado por oportunidad"
        )
    else:
        ws["A2"] = (
            f"Generado: {today} | Perfil: LAASSP, {ncat} categorias | "
            "Licitaciones publicas vigentes"
        )

    counts = {b: 0 for b in buckets}
    for t in ordered:
        counts[bucket_of(t)] += 1
    summary = [(b, f"{b}: {counts[b]}") for b in buckets]
    summary.append((None, f"Total: {len(ordered)}"))
    for i, (bucket, text) in enumerate(summary, start=1):
        cell = ws.cell(row=4, column=i, value=text)
        if bucket:
            bg, fg, _bold = _BUCKET_STYLE[bucket]
            cell.fill = _fill(bg)
            cell.font = Font(bold=True, color=fg)
        else:
            cell.fill = _fill("E5E7EB")
            cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Columns: base through Señal, then (if positioned) the position columns,
    # then Items and Competidores.
    base_headers = [
        "Accion", "No. Procedimiento", "Institucion", "Estado", "Descripcion",
        "Categoria", "Fecha Apertura", "Dias Restantes (cal.)", "Monto Estimado",
        "Banda Historica", "Mediana Hist.", "Tasa Nuevos", "Señal",
    ]
    base_widths = [14, 32, 10, 16, 50, 40, 12, 18, 18, 24, 16, 8, 14]
    pos_headers = ["Posicion", "P Estimada", "Contratos Previos"]
    pos_widths = [16, 12, 16]
    tail_headers = ["Items", "Competidores"]
    tail_widths = [10, 50]
    headers = base_headers + (pos_headers if positioned else []) + tail_headers
    widths = base_widths + (pos_widths if positioned else []) + tail_widths
    col = {name: idx for idx, name in enumerate(headers, start=1)}

    header_row = 6
    for name, idx in col.items():
        cell = ws.cell(row=header_row, column=idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    red_font = Font(color="CC0000")
    amber_font = Font(color="B45309")
    green_font = Font(color="1E7B1E")

    row = header_row + 1
    for t in ordered:
        p = t.primary_intel
        bucket = bucket_of(t)
        days = t.urgency.days_to_deadline
        deadline = t.urgency.deadline.date() if t.urgency.deadline else None
        senal = _senal_es(t.signal)
        low_conf = bool(p and p.has_history and p.confidence == "low")

        values = {
            "Accion": bucket,
            "No. Procedimiento": t.numero_procedimiento,
            "Institucion": t.siglas,
            "Estado": t.entidad_federativa,
            "Descripcion": (t.nombre_procedimiento or "")[:60],
            "Categoria": _categoria_cell(t),
            "Fecha Apertura": deadline,
            "Dias Restantes (cal.)": days,
            "Monto Estimado": _monto_cell(t),
            "Banda Historica": _banda_cell(p),
            "Mediana Hist.": p.price_median if (p and p.price_median is not None) else "S/D",
            "Tasa Nuevos": f"{p.new_entrant_rate:.0%}" if (p and p.new_entrant_rate is not None) else "S/D",
            "Señal": senal + (" *" if low_conf else ""),
            "Items": _items_cell(t),
            "Competidores": _competidores_cell(p),
        }
        if positioned:
            pos = p.position if p else None
            values["Posicion"] = POSITION_ES.get(pos.position_grade, "S/D") if pos else "S/D"
            values["P Estimada"] = f"{pos.p_win_low:.0%}-{pos.p_win_high:.0%}" if pos else "S/D"
            values["Contratos Previos"] = pos.n_prior_wins if pos else 0

        for name, value in values.items():
            cell = ws.cell(row=row, column=col[name], value=value)
            if name == "Fecha Apertura" and deadline is not None:
                cell.number_format = "DD/MM/YYYY"
            elif name == "Mediana Hist." and p and p.price_median is not None:
                cell.number_format = "#,##0"
            elif name == "P Estimada":
                cell.alignment = Alignment(horizontal="center")

        # Conditional styling.
        bg, fg, _bold = _BUCKET_STYLE[bucket]
        acell = ws.cell(row=row, column=col["Accion"])
        acell.fill = _fill(bg)
        acell.font = Font(bold=True, color=fg)

        if days is not None:
            dr = ws.cell(row=row, column=col["Dias Restantes (cal.)"])
            dr.font = red_font if days <= 3 else amber_font if days <= 7 else green_font

        if p and p.new_entrant_rate is not None:
            tn = ws.cell(row=row, column=col["Tasa Nuevos"])
            if p.new_entrant_rate >= 0.30:
                tn.font = green_font
            elif p.new_entrant_rate < 0.15:
                tn.font = red_font

        sbg, sfg, sbold = _SIGNAL_STYLE.get(senal, ("CCCCCC", "000000", False))
        scell = ws.cell(row=row, column=col["Señal"])
        scell.fill = _fill(sbg)
        scell.font = Font(bold=sbold, color=sfg)
        row += 1

    last = max(row - 1, header_row)
    end_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{end_col}{last}"
    ws.freeze_panes = "A7"
    note = ws.cell(row=last + 2, column=1, value="* Señal de baja confianza (pocos datos historicos, n<8)")
    note.font = Font(italic=True, color="6B7280")

    # ---- Sheet 2: Detalle ------------------------------------------------- #
    ws2 = wb.create_sheet("Detalle")
    frame = to_dataframe(shortlist)
    headers = ["Accion"] + [es for _key, es in _DETALLE_COLUMNS]
    for col, name in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font

    for i, t in enumerate(shortlist):
        r = i + 2
        record = frame.iloc[i]
        ws2.cell(row=r, column=1, value=bucket_of(t))
        for col, (key, _es) in enumerate(_DETALLE_COLUMNS, start=2):
            value = record[key]
            if pd.isna(value):
                value = None
            elif hasattr(value, "item"):  # numpy scalar -> python scalar
                value = value.item()
            cell = ws2.cell(row=r, column=col, value=value)
            if key in _DETALLE_MONEY and value is not None:
                cell.number_format = "#,##0"
            elif key in _DETALLE_PERCENT and value is not None:
                cell.number_format = "0%"
            elif key in _DETALLE_RATIO and value is not None:
                cell.number_format = "0.00"
            elif key == "deadline" and value is not None:
                cell.number_format = "DD/MM/YYYY"

    last2 = len(shortlist) + 1
    end_col = get_column_letter(len(headers))
    ws2.auto_filter.ref = f"A1:{end_col}{last2}"
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 50

    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# Prospect (potential-client) list.
# --------------------------------------------------------------------------- #

_PROSPECT_HEADERS = [
    "#", "Empresa", "RFC", "Tamaño", "Contratos", "Licitaciones",
    "Valor Total (MXN)", "Categorias", "Categorias (detalle)", "Compradores",
    "Top Compradores", "Ultimo Año", "Años Activo", "Puntaje",
]
_PROSPECT_WIDTHS = [5, 46, 15, 10, 10, 12, 20, 11, 55, 12, 40, 11, 11, 9]


def default_prospects_path() -> str:
    """Default prospect-list filename, dated for the day it is run."""
    return f"reports/prospectos-veta-{datetime.date.today():%Y-%m-%d}.xlsx"


def write_prospects_xlsx(prospects: list, path: str) -> str:
    """Write the ranked potential-client list to a single-sheet XLSX."""
    check_xlsx_writable(path)
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospectos"

    ws["A1"] = "VETA - Clientes Potenciales"
    ws["A1"].font = Font(bold=True, size=14)
    today = datetime.date.today().strftime("%d/%m/%Y")
    ws["A2"] = (
        f"Generado: {today} | {len(prospects)} empresas | Perfil: proveedores "
        "MIPYME activos que compiten en las categorias objetivo (LAASSP)"
    )
    ws["A3"] = (
        "Nota: ComprasMX no publica contacto/correo. Enriquecer con datos de "
        "contacto antes del envio diario."
    )
    ws["A3"].font = Font(italic=True, color="6B7280")

    header_row = 5
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for col, name in enumerate(_PROSPECT_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
    for col, width in enumerate(_PROSPECT_WIDTHS, start=1):
        ws.column_dimensions[chr(64 + col)].width = width

    row = header_row + 1
    for i, p in enumerate(prospects, start=1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=p.proveedor)
        ws.cell(row=row, column=3, value=p.rfc)
        ws.cell(row=row, column=4, value=p.estratificacion)
        ws.cell(row=row, column=5, value=p.total_contracts)
        ws.cell(row=row, column=6, value=p.licitacion_contracts)
        vcell = ws.cell(row=row, column=7, value=p.total_value)
        vcell.number_format = "#,##0"
        ws.cell(row=row, column=8, value=p.distinct_partidas)
        ws.cell(row=row, column=9, value=", ".join(p.partidas)[:200])
        ws.cell(row=row, column=10, value=p.distinct_buyers)
        ws.cell(row=row, column=11, value=", ".join(p.top_buyers))
        ws.cell(row=row, column=12, value=p.last_year)
        ws.cell(row=row, column=13, value=p.years_active)
        ws.cell(row=row, column=14, value=round(p.score, 1))
        row += 1

    last = max(row - 1, header_row)
    ws.auto_filter.ref = f"A{header_row}:N{last}"
    ws.freeze_panes = f"A{header_row + 1}"
    ws["A1"].alignment = Alignment(vertical="center")

    wb.save(path)
    return path
